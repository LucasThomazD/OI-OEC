import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Carregar os arquivos Excel
arquivo1 = "C:/Users/lucasduarte/Downloads/em desenvolvimento/fisico Financeiro/PPQ Redes - Área 1000 - TA 31 - SALDO.xlsx"
arquivo2 = "C:/Users/lucasduarte/Downloads/em desenvolvimento/fisico Financeiro/PPQ Redes - Área 1000 - TA 33 - SALDO.xlsx"

df1 = pd.read_excel(arquivo1, sheet_name="Novo Relatório")
df2 = pd.read_excel(arquivo2, sheet_name="Novo Relatório")

# Nome da coluna a ser comparada
coluna = "Indice"

# Criar uma cópia do df1 para inserções
df1_modificado = df1.copy()

# Índice de inserção
indice_insercao = 0

# Criar lista para armazenar os índices faltantes
indices_faltantes = []

# Iterar sobre os índices do segundo arquivo
for i, valor in enumerate(df2[coluna]):
    if indice_insercao < len(df1_modificado) and df1_modificado.iloc[indice_insercao][coluna] == valor:
        indice_insercao += 1  # Avança para o próximo índice
    else:
        nova_linha = df2.iloc[i:i+1]  # Pega a linha completa do segundo arquivo
        df1_modificado = pd.concat([df1_modificado.iloc[:indice_insercao], nova_linha, df1_modificado.iloc[indice_insercao:]]).reset_index(drop=True)
        indice_insercao += 1  # Avança após a inserção
        indices_faltantes.append(valor)  # Armazena índice faltante

# Salvar o arquivo modificado
arquivo_modificado = "arquivo1_modificado.xlsx"
df1_modificado.to_excel(arquivo_modificado, index=False)

# Destacar de amarelo os itens faltantes no segundo arquivo
wb = load_workbook(arquivo2)
ws = wb["Novo Relatório"]

fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value in indices_faltantes:  # Verifica se o valor está na lista de faltantes
        row[0].fill = fill  # Aplica o destaque amarelo

wb.save("arquivo2_destacado.xlsx")

print("Comparação concluída! Linhas adicionadas ao primeiro arquivo e itens faltantes destacados em amarelo no segundo.")
